import openpyxl
import re

# 打开Excel文件
wb = openpyxl.load_workbook('科目代码爬虫2.0.xlsx')

# 获取工作表
sheet = wb.active

# 定义字典
result_dict = {}

# 遍历工作表中的每一行
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 获取A列和E列的数据
    key = row[0]
    value = row[4]
    print(type(value))
    # 使用正则表达式提取（）内的内容
    # value_list = re.findall(r'（.*?）', value)
    if value is None or not isinstance(value, str):
        continue
    value_list = re.findall(r'\((.*?)\)', value)
    # 将 key 和 value 存入字典
    if key not in result_dict:
        result_dict[key] = value_list
    else:
        result_dict[key].extend(value_list)

# 输出字典
print(result_dict)

# 首先使用openpyxl库打开Excel文件，然后获取工作表对象。
# 使用iter_rows()方法遍历工作表中的每一行数据。
# 获取A列和E列的数据，并使用正则表达式提取（）内的内容，然后将提取后的值存入字典中。
# 在将数据存入字典之前，检查该key值是否已经存在于字典中，如果存在则将value值扩展到已有的列表中，否则创建新的key-value对。
# 输出最终的字典结果。
